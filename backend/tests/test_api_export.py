"""Export endpoint: txt / csv / xlsx downloads with the full ladder."""

import csv
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.conftest import sample_files, sample_form


def _create_and_match(client: TestClient, name: str = "Kampinos 2 July") -> str:
    analysis_id = str(
        client.post(
            "/api/analyses", data=sample_form(name=name), files=sample_files()
        ).json()["id"]
    )
    client.post(f"/api/analyses/{analysis_id}/match")
    return analysis_id


def test_export_txt(client: TestClient) -> None:
    analysis_id = _create_and_match(client)
    resp = client.get(f"/api/analyses/{analysis_id}/export?format=txt")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/plain")
    assert 'filename="Kampinos 2 July.txt"' in resp.headers["content-disposition"]

    lines = resp.text.strip().split("\n")
    header = lines[0].split("\t")
    assert "Nr" in header
    assert "CO2_Mg_ha_year_co2equiv" in header  # full ladder present
    assert len(lines) == 1 + 4  # header + four spots


def test_export_csv_is_comma_delimited(client: TestClient) -> None:
    analysis_id = _create_and_match(client)
    resp = client.get(f"/api/analyses/{analysis_id}/export?format=csv")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    rows = list(csv.reader(io.StringIO(resp.text)))
    assert rows[0][0] == "Nr"
    assert len(rows) == 1 + 4


def test_export_xlsx_is_valid_workbook(client: TestClient) -> None:
    analysis_id = _create_and_match(client)
    resp = client.get(f"/api/analyses/{analysis_id}/export?format=xlsx")
    assert resp.status_code == 200
    assert "openxmlformats" in resp.headers["content-type"]

    workbook = load_workbook(io.BytesIO(resp.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][0] == "Nr"
    assert len(rows) == 1 + 4
    # Spot 1 (row index 1) has a numeric CO2 umol_m2_s value.
    header = list(rows[0])
    co2_col = header.index("CO2_umol_m2_s")
    assert isinstance(rows[1][co2_col], (int, float))


def test_export_defaults_to_xlsx(client: TestClient) -> None:
    analysis_id = _create_and_match(client)
    resp = client.get(f"/api/analyses/{analysis_id}/export")
    assert resp.status_code == 200
    assert "openxmlformats" in resp.headers["content-type"]


def test_export_unknown_format_422(client: TestClient) -> None:
    analysis_id = _create_and_match(client)
    resp = client.get(f"/api/analyses/{analysis_id}/export?format=pdf")
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "bad_format"


def test_export_unknown_analysis_404(client: TestClient) -> None:
    assert client.get("/api/analyses/nope/export?format=txt").status_code == 404
